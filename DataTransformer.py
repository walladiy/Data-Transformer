import xlwings as xw
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Дата окончания пробного периода
END_DATE = datetime(2025, 2, 15)  

# Функция отвечающая за пробный период
def check_trial_period():
    current_date = datetime.now()
    if current_date > END_DATE:
        return False
    return True

# Функция выбора и конвертации файлов
def choose_and_convert_files():
    root = tk.Tk()
    root.withdraw()  # Скрыть основное окно

    # Открыть диалог выбора нескольких файлов
    file_paths = filedialog.askopenfilenames(title="Выберите файлы", filetypes=[("Excel files", "*.xls")])

    if not file_paths:
        messagebox.showinfo("Информация", "Файлы не выбраны.")
        return []

    # Определяем путь к рабочему столу
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    saved_files = []  # Список для хранения путей сохранённых файлов

    try:
        # Открываем Excel
        app = xw.App(visible=False)
        
        for i, file_path in enumerate(file_paths, start=1):
            try:
                # Открываем файл с помощью xlwings
                wb = xw.Book(file_path)
                sheet = wb.sheets[0]
                data = sheet.used_range.value

                # Преобразуем данные в DataFrame
                df = pd.DataFrame(data[1:], columns=data[0])
                
                # Определяем путь для нового .xlsx файла
                save_path = os.path.join(desktop_path, f"converted_{i}.xlsx")
                
                # Сохранение данных в .xlsx файл
                df.to_excel(save_path, index=False, engine='openpyxl')
                saved_files.append(save_path)  # Добавляем путь к сохранённому файлу в список
                print(f"Файл {file_path} успешно конвертирован и сохранён как {save_path}")

                # Закрываем книгу
                wb.close()
                
            except Exception as e:
                print(f"Ошибка при конвертации файла {file_path}: {e}")

        # Завершаем работу с Excel
        app.quit()
        
    except Exception as e:
        print(f"Ошибка при работе с Excel: {e}")

    return saved_files  # Возвращаем список путей к сохранённым файлам

# Функция переносящая значения "Кол-во" напротив "Наименование"
def copy_value(filePath):
    wb = load_workbook(filePath)
    ws = wb.active

    for row in range(9, ws.max_row + 1, 2):
        ws[f'H{row}'] = ws[f'H{row+1}'].value

    wb.save(filePath)

# Функция удаляющая не нужные ряды и столбцы
def trim_sheet(filePath):
    wb = load_workbook(filePath)
    ws = wb.active

    # Удаляем первые 8 рядов
    for _ in range(8):
        ws.delete_rows(1)
    
    # Удаляем полседние 2 ряда
    for _ in range(2):
        ws.delete_rows(ws.max_row)

    # Удаляем все чётные строки
    for row in range(ws.max_row, 1, -2):
        ws.delete_rows(row)

    # Удаляем ненужные столбцы
    colsList = [9, 7, 6, 5, 4, 2]
    for i in colsList:
        ws.delete_cols(i)

    wb.save(filePath)

# Функция объединяющая файлы в новый
def combine_files(file_list):
    # Создание новой рабочей книги
    wb_comb = Workbook()
    ws_comb = wb_comb.active

    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop') # Путь к рабочему столу
    combined_file = os.path.join(desktop_path, f"combined.xlsx") # Путь объединенного файла

    for file in file_list:
        try:
            # Загрузка текущего файла
            wb = load_workbook(file)
            ws = wb.active

            # Копирование данных из текущего файла
            for row in ws.iter_rows(values_only=True):
                ws_comb.append(row)

        except Exception as e:
            print(f"Ошибка при обработке файла {file}: {e}")

    # Сохранение объединенного файла
    wb_comb.save(combined_file)
    print(f"Файлы успешно объединены и сохранены как {combined_file}")
    return combined_file

# Функция удаляющая промежуточные файлы
def delete_files(file_list):
    for file in file_list:
        try:
            if os.path.exists(file):
                os.remove(file)
                print(f"Файл {file} удалён.")
            else:
                print(f"Файл {file} не существует.")
        except Exception as e:
            print(f"Ошибка при удалении файла {file}: {e}")

# Удаляем "Ед. изм.", укорачиваем строку, форматируем текст и задаем ширину второго столбца
def clean_column(filePath):
    wb = load_workbook(filePath)
    ws = wb.active
    
    # Обрабатываем значения в столбце B (второй столбец)
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=1):  # Столбец B
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            # Разделить строку по 'Ед.изм.'
            parts = cell.value.split('Ед.изм.')
            # Оставить только часть до 'Ед.изм.'
            cell.value = parts[0].strip()

        # Укорачиваем строку
        if cell.value:
            cell.value = cell.value[:25]

    # Форматируем текст
    font = Font(name='Times New Roman', size=10)
    alignment = Alignment(horizontal='left')

    # Устанавливаем ширину второго столбца
    ws.column_dimensions['B'].width = 28  

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    wb.save(filePath)

# Создание интерфейса
class ExcelInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Interface")

        # Определяем размеры экрана
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Определяем размеры окна
        window_width = 400
        window_height = 400

        # Рассчитываем позицию для центрирования
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        # Устанавливаем размеры и позицию окна
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.resizable(False, False)  # Prevent resizing

        # Content display
        self.content_label = tk.Label(root, text="Материальные средства:")
        self.content_label.pack(pady=5)
        self.content_display = tk.Label(root, text="", font=("Arial", 14))
        self.content_display.pack(pady=5)

        # Sheet name input
        self.sheet_name_label = tk.Label(root, text="На что желаете списать:")
        self.sheet_name_label.pack(pady=5)
        self.sheet_name_entry = tk.Entry(root)
        self.sheet_name_entry.pack(pady=5)

        # Bind Enter key to process sheet name
        self.sheet_name_entry.bind("<Return>", self.process_sheet_name)

        # List of created sheets
        self.sheet_list_label = tk.Label(root, text="Созданные листы:")
        self.sheet_list_label.pack(pady=5)
        self.sheet_list_display = tk.Listbox(root, height=10, width=50)
        self.sheet_list_display.pack(pady=5)

        self.file_path = os.path.join(os.path.expanduser("~"), "Desktop", "combined.xlsx")
        self.workbook = None
        self.row_index = 1
        self.is_window_open = True

        # Open file
        self.open_file()

    def open_file(self):
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
            self.update_content_display()
            self.update_sheet_list()
        else:
            messagebox.showerror("File Error", f"File not found: {self.file_path}")

        # Устанавливаем фокус на поле ввода после загрузки окна
        self.root.after(100, self.set_focus_on_entry)

    def set_focus_on_entry(self):
        self.sheet_name_entry.focus_set()  # Устанавливаем фокус на поле ввода

    def update_content_display(self):
        if self.workbook:
            sheet = self.workbook.active
            cell_value = sheet[f'B{self.row_index}'].value
            if cell_value is None:
                self.show_end_of_file_message()
                return
            self.content_display.config(text=cell_value)

    def process_sheet_name(self, event):
        sheet_name = self.sheet_name_entry.get().strip()
        if not sheet_name:
            messagebox.showwarning("Input Error", "Sheet name cannot be empty.")
            return

        if self.workbook:
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.create_sheet(sheet_name)

            # Get the active sheet and copy the row to the new sheet
            active_sheet = self.workbook.active
            row_data = [cell.value for cell in active_sheet[self.row_index]]
            sheet.append(row_data)

            # Устанавливаем ширину второго столбца
            sheet.column_dimensions['B'].width = 28

            # Создаем объект шрифта с требуемым форматированием
            font = Font(name='Times New Roman', size=10)

            # Создаем объект выравнивания для выравнивания по левому краю
            alignment = Alignment(horizontal='left')

            # Применяем шрифт и выравнивание ко всем ячейкам на листе
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
            
            # Save the workbook
            self.workbook.save(self.file_path)

            # Move to the next row in the original sheet
            self.row_index += 1
            self.update_content_display()
            if self.is_window_open:
                self.update_sheet_list()

            # Clear the sheet name entry
            self.sheet_name_entry.delete(0, tk.END)

    def update_sheet_list(self):
        if self.workbook and self.is_window_open:
            self.sheet_list_display.delete(0, tk.END)
            for sheet_name in self.workbook.sheetnames:
                self.sheet_list_display.insert(tk.END, sheet_name)

    def show_end_of_file_message(self):
        # Очистить основное окно
        for widget in self.root.winfo_children():
            widget.pack_forget()

        # Создать сообщение Конец файла
        frame = tk.Frame(self.root)
        frame.pack(expand=True)

        tk.Label(frame, text="Конец файла", font=("Arial", 16)).pack(pady=20)

        button = tk.Button(frame, text="ОК", command=self.close_program, width=20, height=2)
        button.pack(pady=10)

        # Устанавливаем фокус на кнопку после инициализации окна
        self.root.after(100, button.focus_set)  # Устанавливаем фокус на кнопку после инициализации

        # Привязываем клавишу Enter для закрытия программы
        self.root.bind("<Return>", lambda event: self.close_program())

    def close_program(self):
        # Закрытие всех окон и завершение программы
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":

    if not check_trial_period(): # проверка на завершение пробного периода
        exit()  # Завершить выполнение скрипта

    saved_paths = choose_and_convert_files()  # конвертируем файлы и сохраняем к ним пути 

    for file_path in saved_paths:  # обрабатываем файлы
        copy_value(file_path)
        trim_sheet(file_path)

    file_path = combine_files(saved_paths)  # объединяем файлы в новый и записываем путь к нему в переменную
    delete_files(saved_paths)  # удаляем ненужные файлы
    clean_column(file_path)  # удаляем "Ед. изм." и укорачиваем строку
   
    root = tk.Tk()
    app = ExcelInterface(root)
    root.mainloop()
